import os
import openpyxl
from openpyxl.drawing.image import Image
from django.db.models import Avg
from openpyxl.styles import Alignment, Font
from datetime import datetime
from django.conf import settings
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle, Image as PlatypusImage
from reportlab.lib.units import inch, cm
from svglib.svglib import svg2rlg
from reportlab.lib.enums import TA_LEFT, TA_CENTER


def generar_excel_dashboard(data, rango):
    nombre = f"reporte_dashboard_{rango}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    reportes_dir = os.path.join(settings.MEDIA_ROOT, 'reportes')
    os.makedirs(reportes_dir, exist_ok=True)
    ruta = os.path.join(reportes_dir, nombre)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dashboard"

    # --- AÑADIR LOGO ---
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo_emsa.png')
    if os.path.exists(logo_path):
        img = Image(logo_path)
        img.height = 50  # Ajusta la altura
        img.width = 150  # Ajusta el ancho
        ws.add_image(img, 'A1')

    # -------------------------
    #  TITULO
    # -------------------------
    title_cell = ws["C1"]
    title_cell.value = f"Reporte Dashboard ({rango.upper()})"
    title_cell.font = Font(size=18, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells("C1:F2")
    ws.row_dimensions[1].height = 40

    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15

    def _add_section(title, headers, items):
        """Función auxiliar para añadir secciones a la hoja de cálculo."""
        ws.append([])
        header_cell = ws.cell(row=ws.max_row + 1, column=1, value=title)
        header_cell.font = Font(bold=True, size=14)
        ws.append(headers)
        for item_key, item_value in items:
            ws.append([item_key, item_value])

    # ============================
    # KPIs BASE
    # ============================
    kpi_fields = [
        ("Total Tickets", data.get("total_tickets", 0)),
        ("Tickets Hoy", data.get("tickets_hoy", 0)),
        ("Abiertos", data.get("tickets_abiertos", 0)),
        ("En Progreso", data.get("tickets_en_progreso", 0)),
        ("Resueltos", data.get("tickets_resueltos", 0)),
        ("Atrasados", data.get("tickets_atrasados", 0)),
        ("Rating Promedio", round(data.get("promedio_rating", 0), 2)),
        ("Efectividad Global %", round(data.get("efectividad_global", 0), 2)),
        ("Tiempo Prom. Resolución (h)", round(data.get("tiempo_promedio_resolucion", 0), 2)),
    ]
    _add_section("KPIs Generales", ["Métrica", "Valor"], kpi_fields)

    # ============================
    # TICKETS POR CATEGORÍA
    # ============================
    _add_section("Tickets por Categoría", ["Categoría", "Total"], data.get("tickets_por_categoria", {}).items())

    # ============================
    # TICKETS POR PRIORIDAD
    # ============================
    _add_section("Tickets por Prioridad", ["Prioridad", "Total"], data.get("tickets_por_prioridad", {}).items())

    # ============================
    # RATING POR CATEGORÍA
    # ============================
    _add_section("Rating por Categoría", ["Categoría", "Promedio"], data.get("rating_por_categoria", {}).items())

    # ============================
    # RATING POR AGENTE
    # ============================
    _add_section("Rating por Agente", ["Agente", "Promedio"], data.get("rating_por_agente", {}).items())

    wb.save(ruta)
    return ruta


def generar_pdf_dashboard(data, rango):
    """
    Genera un reporte de dashboard en formato PDF.

    Args:
        data (dict): Diccionario con las métricas del dashboard.
        rango (str): El rango de fechas del reporte.
    """
    nombre = f"reporte_dashboard_{rango}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    reportes_dir = os.path.join(settings.MEDIA_ROOT, 'reportes')
    os.makedirs(reportes_dir, exist_ok=True)
    ruta = os.path.join(reportes_dir, nombre)

    doc = SimpleDocTemplate(ruta, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    # --- AÑADIR LOGO Y TÍTULO ---
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo_emsa.png')
    
    header_content = []
    if os.path.exists(logo_path):
        # Usar la clase Image de reportlab para PNG/JPG
        logo_img = PlatypusImage(logo_path, width=4*cm, height=1.5*cm)
        logo_img.hAlign = 'LEFT'
        
        title_style = ParagraphStyle(name='Title', fontSize=24, alignment=TA_CENTER, leading=28)
        title = Paragraph(f"Reporte Dashboard – {rango.upper()}", title_style)
        
        # Usar una tabla para alinear el logo y el título
        header_table = Table([[logo_img, title]], colWidths=[5*cm, doc.width - 5*cm])
        header_table.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'MIDDLE')]))
        elements.append(header_table)
    else:
        elements.append(Paragraph(f"Reporte Dashboard – {rango.upper()}", styles["Title"]))

    elements.append(Spacer(1, 0.3 * inch))

    # =============================
    # KPIs
    # =============================
    kpi_keys = [
        "total_tickets", "tickets_hoy", "tickets_abiertos",
        "tickets_en_progreso", "tickets_resueltos",
        "tickets_atrasados", "promedio_rating"
    ]
    elements.append(Paragraph("KPIs Generales", styles["Heading2"]))
    kpi_data = [
        [key.replace('_', ' ').title(), data.get(key, 'N/A')] for key in kpi_keys
    ]
    
    table = Table(kpi_data, colWidths=[doc.width / 2.0] * 2)
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
    ]))
    elements.append(table)

    elements.append(Spacer(1, 0.3 * inch))

    # =============================
    # GRÁFICO Tickets por Categoría
    # =============================
    elements.append(Paragraph("Tickets por Categoría", styles["Heading2"]))
    
    tickets_por_categoria = data.get("tickets_por_categoria", {})
    if tickets_por_categoria:
        categorias = list(tickets_por_categoria.keys())
        valores = list(tickets_por_categoria.values())

        d = Drawing(400, 240)
        bar = VerticalBarChart()
        bar.x = 50
        bar.y = 50
        bar.height = 150
        bar.width = 300
        bar.data = [valores]
        bar.categoryAxis.categoryNames = [c[:10] for c in categorias] # Acortar nombres largos
        bar.categoryAxis.labels.boxAnchor = 'n'
        d.add(bar)
        elements.append(d)
    else:
        elements.append(Paragraph("No hay datos de tickets por categoría.", styles["BodyText"]))

    elements.append(Spacer(1, 0.3 * inch))

    # =============================
    # GRÁFICO Tickets por Prioridad
    # =============================
    elements.append(Paragraph("Tickets por Prioridad", styles["Heading2"]))
    
    tickets_por_prioridad = data.get("tickets_por_prioridad", {})
    if tickets_por_prioridad:
        prioridades = list(tickets_por_prioridad.keys())
        valores_prio = list(tickets_por_prioridad.values())

        d2 = Drawing(300, 150)
        pie = Pie()
        pie.x = 0
        pie.y = 0
        pie.width = 150
        pie.height = 150
        pie.data = valores_prio
        pie.labels = prioridades
        d2.add(pie)
        elements.append(d2)
    else:
        elements.append(Paragraph("No hay datos de tickets por prioridad.", styles["BodyText"]))

    doc.build(elements)
    return ruta


def generar_excel_tickets(tickets_queryset):
    """
    Genera un reporte de listado de tickets en formato Excel.
    """
    nombre = f"reporte_tickets_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    reportes_dir = os.path.join(settings.MEDIA_ROOT, 'reportes')
    os.makedirs(reportes_dir, exist_ok=True)
    ruta = os.path.join(reportes_dir, nombre)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Listado de Tickets"

    # --- AÑADIR LOGO ---
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo_emsa.png')
    if os.path.exists(logo_path):
        img = Image(logo_path)
        img.height = 50
        img.width = 150
        ws.add_image(img, 'A1')

    # --- Título ---
    title_cell = ws["C1"]
    title_cell.value = "Reporte de Tickets"
    title_cell.font = Font(size=18, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells("C1:J2")
    ws.row_dimensions[1].height = 40

    # --- Encabezados ---
    headers = [
        "ID", "Título", "Estado", "Prioridad", "Categoría",
        "Solicitante", "Agente", "Fecha Creación", "Fecha Cierre", "Rating"
    ]
    ws.append([])
    ws.append(headers)
    for cell in ws[ws.max_row]: # Corregido: aplicar estilo a la última fila (la de los encabezados)
        cell.font = Font(bold=True)

    # --- Datos ---
    for ticket in tickets_queryset:
        ws.append([
            ticket.id,
            ticket.titulo,
            ticket.get_estado_display(),
            ticket.prioridad if ticket.prioridad else "N/A",
            ticket.categoria_principal.nombre if ticket.categoria_principal else "N/A",
            ticket.solicitante.username if ticket.solicitante else "N/A",
            ticket.agente.username if ticket.agente else "Sin asignar",
            ticket.fecha_creacion.strftime("%Y-%m-%d %H:%M") if ticket.fecha_creacion else "",
            ticket.fecha_cierre.strftime("%Y-%m-%d %H:%M") if ticket.fecha_cierre else "",
            ticket.rating if ticket.rating is not None else "Sin calificar"
        ])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[2].column_letter # Usar una fila que no esté combinada (fila 3 de encabezados)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(ruta)
    return ruta


def generar_pdf_tickets(tickets_queryset):
    """
    Genera un reporte de listado de tickets en formato PDF.
    """
    nombre = f"reporte_tickets_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    reportes_dir = os.path.join(settings.MEDIA_ROOT, 'reportes')
    os.makedirs(reportes_dir, exist_ok=True)
    ruta = os.path.join(reportes_dir, nombre)

    doc = SimpleDocTemplate(ruta, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    elements = []

    # --- AÑADIR LOGO Y TÍTULO ---
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo_emsa.png')
    if os.path.exists(logo_path):
        logo_img = PlatypusImage(logo_path, width=4*cm, height=1.5*cm)
        logo_img.hAlign = 'LEFT'
        
        title_style = ParagraphStyle(name='Title', fontSize=24, alignment=TA_CENTER, leading=28)
        title = Paragraph("Reporte de Tickets", title_style)
        
        header_table = Table([[logo_img, title]], colWidths=[5*cm, doc.width - 5*cm])
        header_table.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'MIDDLE')]))
        elements.append(header_table)
    else:
        elements.append(Paragraph("Reporte de Tickets", styles["Title"]))

    elements.append(Spacer(1, 0.3 * inch))

    # --- Datos de la tabla ---
    data = [
        ["ID", "Título", "Estado", "Agente", "Fecha Creación"] # Encabezados
    ]
    for ticket in tickets_queryset:
        data.append([
            ticket.id,
            Paragraph(ticket.titulo, styles['BodyText']),
            ticket.get_estado_display(),
            ticket.agente.username if ticket.agente else "N/A",
            ticket.fecha_creacion.strftime("%d/%m/%Y")
        ])

    table = Table(data, colWidths=[0.5*inch, 2.5*inch, 1*inch, 1.5*inch, 1.2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    elements.append(table)
    doc.build(elements)
    return ruta


def generar_excel_categorias(category_queryset):
    """
    Genera un reporte de métricas por categoría en formato Excel.
    """
    nombre = f"reporte_categorias_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    reportes_dir = os.path.join(settings.MEDIA_ROOT, 'reportes')
    os.makedirs(reportes_dir, exist_ok=True)
    ruta = os.path.join(reportes_dir, nombre)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Métricas por Categoría"

    # --- AÑADIR LOGO ---
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo_emsa.png')
    if os.path.exists(logo_path):
        img = Image(logo_path)
        img.height = 50
        img.width = 150
        ws.add_image(img, 'A1')

    # --- Título ---
    title_cell = ws["C1"]
    title_cell.value = "Reporte por Categorías"
    title_cell.font = Font(size=18, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells("C1:F2")
    ws.row_dimensions[1].height = 40

    # --- Encabezados ---
    headers = [
        "Categoría", "Total Tickets", "Tickets Abiertos",
        "Tickets Resueltos", "Rating Promedio", "Tiempo Prom. Resolución (h)"
    ]
    ws.append([])
    ws.append(headers)
    for cell in ws[3]:
        cell.font = Font(bold=True)

    # --- Datos ---
    for cat in category_queryset:
        ws.append([
            cat.nombre,
            cat.total_tickets,
            cat.tickets_abiertos,
            cat.tickets_resueltos,
            round(cat.rating_promedio or 0, 2),
            round(cat.tiempo_promedio_resolucion_horas or 0, 2)
        ])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[2].column_letter # Usar una fila que no esté combinada (fila 3 de encabezados)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(ruta)
    return ruta


def generar_pdf_categorias(category_queryset):
    """
    Genera un reporte de métricas por categoría en formato PDF.
    """
    nombre = f"reporte_categorias_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    reportes_dir = os.path.join(settings.MEDIA_ROOT, 'reportes')
    os.makedirs(reportes_dir, exist_ok=True)
    ruta = os.path.join(reportes_dir, nombre)

    doc = SimpleDocTemplate(ruta, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    # --- AÑADIR LOGO Y TÍTULO ---
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo_emsa.png')
    if os.path.exists(logo_path):
        logo_img = PlatypusImage(logo_path, width=4*cm, height=1.5*cm)
        logo_img.hAlign = 'LEFT'
        
        title_style = ParagraphStyle(name='Title', fontSize=24, alignment=TA_CENTER, leading=28)
        title = Paragraph("Reporte por Categorías", title_style)
        
        header_table = Table([[logo_img, title]], colWidths=[5*cm, doc.width - 5*cm])
        header_table.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'MIDDLE')]))
        elements.append(header_table)
    else:
        elements.append(Paragraph("Reporte por Categorías", styles["Title"]))

    elements.append(Spacer(1, 0.3 * inch))

    data = [["Categoría", "Total Tickets", "Rating Promedio", "Tiempo Prom. Res. (h)"]]
    for cat in category_queryset:
        data.append([cat.nombre, cat.total_tickets, round(cat.rating_promedio or 0, 2), round(cat.tiempo_promedio_resolucion_horas or 0, 2)])

    table = Table(data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch, 1.8*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.teal),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    elements.append(table)
    doc.build(elements)
    return ruta


def generar_excel_usuarios(users_queryset):
    """
    Genera un reporte de listado de usuarios en formato Excel.
    """
    nombre = f"reporte_usuarios_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    reportes_dir = os.path.join(settings.MEDIA_ROOT, 'reportes')
    os.makedirs(reportes_dir, exist_ok=True)
    ruta = os.path.join(reportes_dir, nombre)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Listado de Usuarios"

    # --- AÑADIR LOGO ---
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo_emsa.png')
    if os.path.exists(logo_path):
        img = Image(logo_path)
        img.height = 50
        img.width = 150
        ws.add_image(img, 'A1')

    # --- Título ---
    title_cell = ws["C1"]
    title_cell.value = "Reporte de Usuarios"
    title_cell.font = Font(size=18, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells("C1:H2")
    ws.row_dimensions[1].height = 40

    # --- Encabezados ---
    headers = [
        "ID", "Username", "Nombre Completo", "Email", "Rol",
        "Activo", "Fecha Registro", "Último Login"
    ]
    ws.append([])
    ws.append(headers)
    for cell in ws[3]:
        cell.font = Font(bold=True)

    # --- Datos ---
    for user in users_queryset:
        nombre_completo = f"{user.first_name} {user.last_name}".strip() or user.username
        ws.append([
            user.id,
            user.username,
            nombre_completo,
            user.email,
            user.rol.nombre_visible if user.rol else "Sin rol",
            "Sí" if user.is_active else "No",
            user.date_joined.strftime("%Y-%m-%d %H:%M") if user.date_joined else "",
            user.last_login.strftime("%Y-%m-%d %H:%M") if user.last_login else "Nunca",
        ])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[2].column_letter # Usar una fila que no esté combinada (fila 3 de encabezados)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(ruta)
    return ruta


def generar_pdf_usuarios(users_queryset):
    """
    Genera un reporte de listado de usuarios en formato PDF.
    """
    nombre = f"reporte_usuarios_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    reportes_dir = os.path.join(settings.MEDIA_ROOT, 'reportes')
    os.makedirs(reportes_dir, exist_ok=True)
    ruta = os.path.join(reportes_dir, nombre)

    doc = SimpleDocTemplate(ruta, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    # --- AÑADIR LOGO Y TÍTULO ---
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo_emsa.png')
    if os.path.exists(logo_path):
        logo_img = PlatypusImage(logo_path, width=4*cm, height=1.5*cm)
        logo_img.hAlign = 'LEFT'
        
        title_style = ParagraphStyle(name='Title', fontSize=24, alignment=TA_CENTER, leading=28)
        title = Paragraph("Reporte de Usuarios", title_style)
        
        header_table = Table([[logo_img, title]], colWidths=[5*cm, doc.width - 5*cm])
        header_table.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'MIDDLE')]))
        elements.append(header_table)
    else:
        elements.append(Paragraph("Reporte de Usuarios", styles["Title"]))

    elements.append(Spacer(1, 0.3 * inch))

    data = [["ID", "Username", "Email", "Rol", "Activo"]]
    for user in users_queryset:
        data.append([
            user.id, user.username, user.email, 
            user.rol.nombre_visible if user.rol else "Sin rol", 
            "Sí" if user.is_active else "No"
        ])

    table = Table(data, colWidths=[0.5*inch, 2*inch, 2.5*inch, 1.5*inch, 0.8*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    elements.append(table)
    doc.build(elements)
    return ruta


def generar_excel_rendimiento(performance_queryset):
    """
    Genera un reporte de rendimiento de agentes en formato Excel.
    """
    nombre = f"reporte_rendimiento_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    reportes_dir = os.path.join(settings.MEDIA_ROOT, 'reportes')
    os.makedirs(reportes_dir, exist_ok=True)
    ruta = os.path.join(reportes_dir, nombre)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rendimiento de Agentes"

    # --- AÑADIR LOGO ---
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo_emsa.png')
    if os.path.exists(logo_path):
        img = Image(logo_path)
        img.height = 50
        img.width = 150
        ws.add_image(img, 'A1')

    # --- Título ---
    title_cell = ws["C1"]
    title_cell.value = "Reporte de Rendimiento de Agentes"
    title_cell.font = Font(size=18, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells("C1:H2")
    ws.row_dimensions[1].height = 40

    # --- Encabezados ---
    headers = [
        "Agente", "Rol", "Tickets Asignados", "Tickets Resueltos",
        "Tiempo Prom. Resolución (h)", "Efectividad (%)", "Rating Promedio", "Actualizado en"
    ]
    ws.append([])
    ws.append(headers)
    for cell in ws[3]:
        cell.font = Font(bold=True)

    # --- Datos ---
    for perf in performance_queryset:
        # Calcular el rating promedio aquí, ya que no es un campo directo del modelo
        rating_avg = perf.agente.tickets_asignados.filter(rating__isnull=False).aggregate(Avg('rating'))['rating__avg'] or 0
        ws.append([
            perf.agente.username,
            perf.agente.rol.nombre_visible if perf.agente.rol else "Sin rol",
            perf.tickets_asignados,
            perf.tickets_resueltos,
            round(perf.tiempo_promedio_resolucion, 2),
            round(perf.efectividad, 2),
            round(rating_avg, 2),
            perf.actualizado_en.strftime("%Y-%m-%d %H:%M") if perf.actualizado_en else "",
        ])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[2].column_letter # Usar una fila que no esté combinada (fila 3 de encabezados)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(ruta)
    return ruta


def generar_pdf_rendimiento(performance_queryset):
    """
    Genera un reporte de rendimiento de agentes en formato PDF.
    """
    nombre = f"reporte_rendimiento_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    reportes_dir = os.path.join(settings.MEDIA_ROOT, 'reportes')
    os.makedirs(reportes_dir, exist_ok=True)
    ruta = os.path.join(reportes_dir, nombre)

    doc = SimpleDocTemplate(ruta, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    # --- AÑADIR LOGO Y TÍTULO ---
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo_emsa.png')
    if os.path.exists(logo_path):
        logo_img = PlatypusImage(logo_path, width=4*cm, height=1.5*cm)
        logo_img.hAlign = 'LEFT'
        
        title_style = ParagraphStyle(name='Title', fontSize=24, alignment=TA_CENTER, leading=28)
        title = Paragraph("Reporte de Rendimiento de Agentes", title_style)
        
        header_table = Table([[logo_img, title]], colWidths=[5*cm, doc.width - 5*cm])
        header_table.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'MIDDLE')]))
        elements.append(header_table)
    else:
        elements.append(Paragraph("Reporte de Rendimiento de Agentes", styles["Title"]))

    elements.append(Spacer(1, 0.3 * inch))

    data = [["Agente", "Tickets Resueltos", "Efectividad (%)", "Rating Promedio"]]
    for perf in performance_queryset:
        # Calcular el rating promedio aquí, ya que no es un campo directo del modelo
        rating_avg = perf.agente.tickets_asignados.filter(rating__isnull=False).aggregate(Avg('rating'))['rating__avg'] or 0
        data.append([
            perf.agente.username, perf.tickets_resueltos, 
            f"{round(perf.efectividad, 2)}%", round(rating_avg, 2)
        ])

    table = Table(data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.purple),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    elements.append(table)
    doc.build(elements)
    return ruta
